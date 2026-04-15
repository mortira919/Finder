"""
server.py — FastAPI веб-сервер для WorkWork Studio Lead Finder
"""

import os
import uuid
import json
import asyncio
import threading
import time
from typing import Dict, Optional
from pathlib import Path

import io
from fastapi import FastAPI, Request, UploadFile, File
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
from database import init_db, save_search, get_history, get_filepath_by_job

load_dotenv()

app = FastAPI(title="WorkWork Lead Finder")
init_db()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Хранилище активных задач
jobs: Dict[str, dict] = {}


def run_search(job_id: str, city: str, country: str, categories: list, api_key: str):
    """Запускает поиск в отдельном потоке."""
    import sys
    sys.path.insert(0, os.path.dirname(__file__))

    os.environ["GOOGLE_MAPS_API_KEY"] = api_key
    os.environ["SEARCH_CITY"] = city
    os.environ["COUNTRY"] = country
    os.environ["SEARCH_CATEGORIES"] = ",".join(categories)
    os.environ["MAX_RESULTS"] = "20"

    job = jobs[job_id]

    try:
        from finder import find_companies
        from analyzer import analyze_website
        from reporter import generate_report

        # Шаг 1 — поиск по категориям с реальным прогрессом
        job["status"] = "searching"
        job["message"] = f"Ищем компании в {city}..."
        job["progress"] = 3

        total_cats = len(categories)

        def search_progress(done, total, message):
            # 3% → 18% равномерно по категориям
            job["progress"] = 3 + int(done / total * 15)
            job["message"] = message

        companies = find_companies(categories, on_progress=search_progress)
        job["total"] = len(companies)
        job["message"] = f"Найдено {len(companies)} компаний. Начинаем анализ сайтов..."
        job["progress"] = 18

        # Шаг 2 — анализ сайтов с реальным прогрессом
        job["status"] = "analyzing"
        total_companies = len(companies)
        for i, company in enumerate(companies):
            name = company.get("name", "...")
            website = company.get("website", "")

            if website:
                job["message"] = f"Анализируем сайт: {name} ({i+1}/{total_companies})"
                company["site_analysis"] = analyze_website(website)
            else:
                job["message"] = f"Нет сайта: {name} ({i+1}/{total_companies})"
                company["site_analysis"] = analyze_website("")

            company["app_check"] = {
                "has_mobile_app": False,
                "app_status": "Не проверялось",
                "ios_app": False, "android_app": False,
                "ios_link": "", "android_link": ""
            }

            # 18% → 93% равномерно по компаниям
            job["progress"] = 18 + int((i + 1) / total_companies * 75)

        # Шаг 3
        job["status"] = "generating"
        job["message"] = "Генерируем отчёт..."
        job["progress"] = 95

        import re as _re
        from reporter import _is_hot_lead, _get_priority_label, _get_email_subject, _generate_email_body
        hot = [c for c in companies if _is_hot_lead(c)]

        # Собираем карточки лидов для UI
        leads_data = []
        for c in hot:
            site = c.get("site_analysis", {})
            socials = site.get("socials", {})
            phone = c.get("phone", "")

            wa = socials.get("whatsapp", "")
            if not wa and phone:
                digits = _re.sub(r"[^\d]", "", phone)
                if digits:
                    wa = f"https://wa.me/{digits}"
            elif wa and not wa.startswith("http"):
                wa = "https://" + wa

            ig = socials.get("instagram", "")
            if ig and not ig.startswith("http"):
                ig = "https://" + ig

            tg = socials.get("telegram", "")
            if tg and not tg.startswith("http"):
                tg = "https://" + tg

            leads_data.append({
                "name": c.get("name", ""),
                "category": c.get("category", ""),
                "phone": phone,
                "whatsapp": wa,
                "instagram": ig,
                "telegram": tg,
                "website": c.get("website", ""),
                "rating": c.get("rating", ""),
                "reviews": c.get("reviews_count", 0),
                "priority": _get_priority_label(c),
                "verdict": site.get("verdict", "Нет сайта"),
                "email_subject": _get_email_subject(c),
                "email_body": _generate_email_body(c),
            })

        job["leads"] = leads_data

        filepath = generate_report(companies, country=country)
        no_site = sum(1 for c in companies if not c.get("website"))

        job["status"] = "done"
        job["progress"] = 100
        job["file"] = filepath
        job["hot_count"] = len(hot)
        job["total_count"] = len(companies)
        job["no_site_count"] = no_site
        job["message"] = f"Готово! Найдено {len(hot)} горячих лидов."

        # Сохраняем в историю
        save_search(
            job_id=job_id, city=city, country=country, categories=categories,
            hot_count=len(hot), total_count=len(companies),
            no_site_count=no_site, filepath=filepath
        )

    except Exception as e:
        job["status"] = "error"
        job["message"] = f"Ошибка: {str(e)}"
        job["progress"] = 0


@app.get("/api/ping")
async def ping():
    return {"ok": True}


@app.get("/", response_class=HTMLResponse)
async def index():
    html_path = Path(__file__).parent / "templates" / "index.html"
    return HTMLResponse(html_path.read_text(encoding="utf-8"))


@app.post("/api/search")
async def start_search(request: Request):
    data = await request.json()
    job_id = str(uuid.uuid4())

    jobs[job_id] = {
        "status": "starting",
        "progress": 0,
        "message": "Запуск...",
        "file": None,
        "hot_count": 0,
        "total_count": 0,
        "no_site_count": 0,
    }

    thread = threading.Thread(
        target=run_search,
        args=(
            job_id,
            data.get("city", "Алматы"),
            data.get("country", "KZ"),
            data.get("categories", ["ресторан", "салон красоты", "стоматология",
                                    "фитнес клуб", "автосервис", "отель", "аптека"]),
            os.getenv("GOOGLE_MAPS_API_KEY", ""),
        ),
        daemon=True
    )
    thread.start()

    return {"job_id": job_id}


@app.get("/api/status/{job_id}")
async def get_status(job_id: str):
    job = jobs.get(job_id)
    if not job:
        return {"status": "not_found"}
    return job


@app.post("/api/import")
async def import_excel(file: UploadFile = File(...)):
    """Парсит загруженный Excel-файл и возвращает список лидов."""
    from openpyxl import load_workbook

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        return JSONResponse({"error": f"Не удалось открыть файл: {e}"}, status_code=400)

    # Ищем листы с лидами в порядке приоритета
    preferred = ["Все горячие лиды", "Горячие лиды", "Менеджер 1", "Менеджер 2"]
    ws = next((wb[n] for n in preferred if n in wb.sheetnames), wb.active)

    leads = []
    headers = None

    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(c).strip() if c else "" for c in row]
            continue
        if not any(row):
            continue

        def g(col):
            try:
                return str(row[headers.index(col)] or "").strip()
            except (ValueError, IndexError):
                return ""

        phone = g("Телефон")
        wa    = g("WhatsApp")
        ig    = g("Instagram")
        tg    = g("Telegram")

        # Приводим к полным URL
        if wa and not wa.startswith("http"): wa = "https://" + wa
        if ig and not ig.startswith("http"): ig = "https://" + ig
        if tg and not tg.startswith("http"): tg = "https://" + tg

        name = g("Название")
        if not name:
            continue

        leads.append({
            "name":          name,
            "category":      g("Категория"),
            "phone":         phone,
            "whatsapp":      wa,
            "instagram":     ig,
            "telegram":      tg,
            "website":       g("Сайт"),
            "rating":        g("Рейтинг"),
            "reviews":       g("Отзывов"),
            "priority":      g("Приоритет") or "📋 НИЗКИЙ",
            "verdict":       g("Статус сайта"),
            "email_subject": g("Шаблон письма"),
            "email_body":    "",
        })

    return {"leads": leads, "total": len(leads), "sheet": ws.title}


@app.get("/api/history")
async def history():
    return get_history()


@app.get("/api/download/{job_id}")
async def download(job_id: str):
    # Сначала ищем в памяти (текущая сессия)
    job = jobs.get(job_id)
    filepath = job["file"] if job and job.get("file") else None

    # Если нет в памяти — смотрим в БД (перезапуск сервера)
    if not filepath:
        filepath = get_filepath_by_job(job_id)

    if not filepath or not os.path.exists(filepath):
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Файл не найден. Возможно, он был удалён."}, status_code=404)

    filename = os.path.basename(filepath)
    return FileResponse(
        filepath,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("server:app", host="0.0.0.0", port=8000, reload=False)
