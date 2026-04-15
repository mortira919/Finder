"""
reporter.py — Генерирует Excel-отчёт и шаблоны холодных писем
"""

import os
import re
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from colorama import Fore, Style

OUTPUT_DIR = "output"


def generate_report(companies: List[Dict], country: str = "RU") -> str:
    """
    Генерирует Excel-файл с результатами анализа.
    country="KZ" → делит горячие лиды между 2 менеджерами.
    country="RU" → один общий лист горячих лидов.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(OUTPUT_DIR, f"leads_{timestamp}.xlsx")

    wb = Workbook()
    hot_leads = [c for c in companies if _is_hot_lead(c)]

    if country == "KZ":
        # КЗ: делим горячих лидов между 2 менеджерами
        mid = len(hot_leads) // 2
        manager1_leads = hot_leads[:mid]
        manager2_leads = hot_leads[mid:]

        for c in manager1_leads:
            c["manager"] = "Менеджер 1"
        for c in manager2_leads:
            c["manager"] = "Менеджер 2"

        ws_all = wb.active
        ws_all.title = "Все горячие лиды"
        _fill_leads_sheet(ws_all, hot_leads, use_manager_field=True)

        ws_m1 = wb.create_sheet("Менеджер 1")
        _fill_leads_sheet(ws_m1, manager1_leads, manager_name="Менеджер 1")

        ws_m2 = wb.create_sheet("Менеджер 2")
        _fill_leads_sheet(ws_m2, manager2_leads, manager_name="Менеджер 2")
    else:
        # РФ: один лист горячих лидов
        ws_all = wb.active
        ws_all.title = "Горячие лиды"
        _fill_leads_sheet(ws_all, hot_leads)

    # Шаблоны писем + Статистика
    ws_emails = wb.create_sheet("Шаблоны писем")
    _fill_email_templates(ws_emails, hot_leads)

    ws_stats = wb.create_sheet("Статистика")
    _fill_stats_sheet(ws_stats, companies)

    wb.save(filepath)
    return filepath


def _fill_leads_sheet(ws, companies: List[Dict], manager_name: str = "", use_manager_field: bool = False):
    """Заполняет лист с лидами."""
    headers = [
        "Менеджер", "Приоритет", "Название", "Категория", "Адрес",
        "Телефон", "WhatsApp", "Сайт", "Instagram", "VK", "Telegram",
        "Рейтинг", "Отзывов", "Статус сайта", "Скор (0-100)",
        "Возможность", "Шаблон письма"
    ]

    # Стили заголовков
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 35

    # Сортируем: сначала горячие (низкий скор = больший приоритет)
    sorted_companies = sorted(
        companies,
        key=lambda c: (
            not _is_hot_lead(c),
            c.get("site_analysis", {}).get("score", 100)
        )
    )

    for row_idx, company in enumerate(sorted_companies, 2):
        site = company.get("site_analysis", {})
        apps = company.get("app_check", {})
        priority = _get_priority_label(company)

        mgr = company.get("manager", manager_name) if use_manager_field else manager_name
        row_fill_color = _get_manager_color(mgr) or _get_row_color(priority)

        socials = site.get("socials", {})
        phone = company.get("phone", "")
        wa = socials.get("whatsapp", "")
        if not wa and phone:
            digits = re.sub(r"[^\d]", "", phone)
            if digits:
                wa = f"wa.me/{digits}"

        row_data = [
            mgr,
            priority,
            company.get("name", ""),
            company.get("category", ""),
            company.get("address", ""),
            phone,
            wa,
            company.get("website", ""),
            socials.get("instagram", ""),
            socials.get("vk", ""),
            socials.get("telegram", ""),
            company.get("rating", ""),
            company.get("reviews_count", ""),
            site.get("verdict", ""),
            site.get("score", ""),
            _get_opportunity_summary(company),
            _get_email_subject(company),
        ]

        row_fill_color = _get_manager_color(manager_name) or _get_row_color(priority)

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
            if row_fill_color:
                cell.fill = PatternFill(
                    start_color=row_fill_color,
                    end_color=row_fill_color,
                    fill_type="solid"
                )

        ws.row_dimensions[row_idx].height = 25

    # Ширина колонок
    column_widths = [14, 12, 28, 16, 32, 16, 22, 30, 28, 20, 18, 8, 8, 20, 10, 38, 45]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"


def _fill_email_templates(ws, companies: List[Dict]):
    """Лист с готовыми персонализированными письмами."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 100

    headers = ["Компания", "Тема письма", "Текст письма"]
    header_fill = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row_idx, company in enumerate(companies, 2):
        subject = _get_email_subject(company)
        body = _generate_email_body(company)

        ws.cell(row=row_idx, column=1, value=company.get("name", ""))
        ws.cell(row=row_idx, column=2, value=subject)
        cell = ws.cell(row=row_idx, column=3, value=body)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_idx].height = 120


def _fill_stats_sheet(ws, companies: List[Dict]):
    """Лист со статистикой."""
    total = len(companies)
    no_website = sum(1 for c in companies if not c.get("website"))
    no_app = sum(1 for c in companies if not c.get("app_check", {}).get("has_mobile_app"))
    outdated_site = sum(
        1 for c in companies
        if c.get("site_analysis", {}).get("score", 100) < 45
    )
    hot = sum(1 for c in companies if _is_hot_lead(c))

    stats = [
        ("📊 Общая статистика", ""),
        ("", ""),
        ("Всего компаний проанализировано", total),
        ("Нет сайта", no_website),
        ("Устаревший сайт (скор < 45)", outdated_site),
        ("Нет мобильного приложения", no_app),
        ("🔥 Горячих лидов", hot),
        ("", ""),
        ("Конверсионный потенциал", f"{hot}/{total} = {int(hot/total*100) if total else 0}%"),
    ]

    header_fill = PatternFill(start_color="0f3460", end_color="0f3460", fill_type="solid")

    for row_idx, (label, value) in enumerate(stats, 1):
        cell_label = ws.cell(row=row_idx, column=1, value=label)
        cell_value = ws.cell(row=row_idx, column=2, value=value)

        if row_idx == 1:
            cell_label.font = Font(bold=True, size=14, color="FFFFFF")
            cell_label.fill = header_fill

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20


def _is_hot_lead(company: Dict) -> bool:
    """Горячий лид = нет сайта ИЛИ плохой сайт И нет приложения."""
    site_score = company.get("site_analysis", {}).get("score", 50)
    has_app = company.get("app_check", {}).get("has_mobile_app", True)
    has_website = bool(company.get("website"))

    return (not has_website or site_score < 45) and not has_app


def _get_priority_label(company: Dict) -> str:
    site_score = company.get("site_analysis", {}).get("score", 50)
    has_app = company.get("app_check", {}).get("has_mobile_app", True)
    has_website = bool(company.get("website"))

    if not has_website and not has_app:
        return "🔥 ВЫСОКИЙ"
    elif site_score < 30 and not has_app:
        return "🔥 ВЫСОКИЙ"
    elif site_score < 50 or not has_app:
        return "⚡ СРЕДНИЙ"
    else:
        return "📋 НИЗКИЙ"


def _get_row_color(priority: str) -> Optional[str]:
    if "ВЫСОКИЙ" in priority:
        return "fff3cd"  # Жёлтый
    elif "СРЕДНИЙ" in priority:
        return "d1ecf1"  # Голубой
    return None


def _get_manager_color(manager_name: str) -> Optional[str]:
    if manager_name == "Менеджер 1":
        return "dce8f5"  # Голубой — Менеджер 1
    elif manager_name == "Менеджер 2":
        return "e8f5dc"  # Зелёный — Менеджер 2
    return None


def _get_opportunity_summary(company: Dict) -> str:
    parts = []
    if not company.get("website"):
        parts.append("Нет сайта")
    else:
        opp = company.get("site_analysis", {}).get("opportunity", "")
        if opp and "в порядке" not in opp.lower():
            parts.append(opp)

    app_status = company.get("app_check", {}).get("app_status", "")
    if app_status == "Нет приложения":
        parts.append("Нет мобильного приложения")

    return " | ".join(parts) if parts else "Уже хорошо"


def _get_email_subject(company: Dict) -> str:
    name = company.get("name", "")
    has_website = bool(company.get("website"))
    site_score = company.get("site_analysis", {}).get("score", 50)

    if not has_website:
        return f"Мобильное приложение для {name} — ваши клиенты уже в смартфонах"
    elif site_score < 45:
        return f"Приложение и новый сайт для {name} — больше клиентов каждый день"
    else:
        return f"Мобильное приложение для {name} — автоматизируйте продажи"


def _generate_email_body(company: Dict) -> str:
    name = company.get("name", "компании")
    category = company.get("category", "бизнеса")
    has_website = bool(company.get("website"))
    site_score = company.get("site_analysis", {}).get("score", 50)
    site_issues = company.get("site_analysis", {}).get("opportunity", "")

    greeting = f"Здравствуйте!\n\nМеня зовут [Имя], я из команды WorkWork Studio — мы разрабатываем мобильные приложения и сайты для бизнеса.\n\n"

    if not has_website:
        pain = (
            f"Заметили, что у {name} пока нет ни сайта, ни мобильного приложения. "
            f"Сегодня 80% клиентов ищут услуги через смартфон — без цифрового присутствия "
            f"вы теряете их каждый день.\n\n"
            f"Мы можем разработать для вас:\n"
            f"• Мобильное приложение (iOS + Android) — онлайн-запись, оплата, уведомления\n"
            f"• Современный сайт — в топе Google и 2ГИС, заявки 24/7\n"
            f"• Всё под ключ с гарантией 1 год\n\n"
        )
    elif site_score < 45:
        pain = (
            f"Сайт {name} есть, но он устарел — {site_issues}. "
            f"Клиенты уходят ещё до того, как позвонили.\n\n"
            f"Мы предлагаем:\n"
            f"• Мобильное приложение (iOS + Android) — повторные продажи и лояльность\n"
            f"• Обновление сайта — скорость, адаптивность, конверсия\n"
            f"• Push-уведомления, онлайн-запись, оплата прямо в приложении\n\n"
        )
    else:
        pain = (
            f"У {name} хороший сайт — но мобильного приложения нет. "
            f"80% ваших клиентов сидят в смартфонах и предпочитают приложения браузеру.\n\n"
            f"Приложение от WorkWork Studio даст вам:\n"
            f"• Повторные визиты +40% через push-уведомления\n"
            f"• Онлайн-запись и оплату прямо в приложении\n"
            f"• Программу лояльности и личный кабинет клиента\n\n"
        )

    cta = (
        f"Готовы показать конкретные решения под ваш бизнес — "
        f"это займёт 15 минут.\n\n"
        f"Удобно ли созвониться на этой неделе?\n\n"
        f"С уважением,\n"
        f"Команда WorkWork Studio\n"
        f"https://workworkstudio.pro"
    )

    return greeting + pain + cta


